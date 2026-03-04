import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os
import re

# --- CẤU HÌNH ---
st.set_page_config(page_title="Hệ Thống Ôn Tập FIC 2026", layout="wide")

# Khởi tạo Session State
if 'high_scores' not in st.session_state:
    st.session_state.high_scores = {} 
if 'submitted_lessons' not in st.session_state:
    st.session_state.submitted_lessons = set() 

# Sắp xếp tự nhiên (Bài 1 -> Bài 2 -> Bài 10)
def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', s)]

def get_file_path():
    files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~$')]
    return files[0] if files else "De_on_tap.xlsx"

@st.cache_data
def load_all_data(file_path):
    if not os.path.exists(file_path): return None
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        all_lessons = {}
        last_lesson_no, last_lesson_name = "", ""
        
        for row_idx in range(1, ws.max_row + 1):
            val_a = ws.cell(row=row_idx, column=1).value 
            val_b = ws.cell(row=row_idx, column=2).value 
            if val_a and "Bài" in str(val_a):
                last_lesson_no = str(val_a).strip()
                last_lesson_name = str(val_b).strip() if val_b else ""
            
            if not last_lesson_no: continue
            current_key = f"{last_lesson_no}: {last_lesson_name}"
            
            question = ws.cell(row=row_idx, column=4).value
            if not question or str(question).strip() in ["Câu hỏi", "Thứ tự", "STT"]: continue
            
            options, correct_options = [], []
            for col_idx in range(6, 12): # Cột F đến K
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    val_str = str(cell.value).strip()
                    options.append(val_str)
                    fill = cell.fill
                    if fill and fill.start_color and fill.start_color.index not in ("00000000", "FFFFFFFF", "0", "None"):
                        correct_options.append(val_str)

            if options:
                if current_key not in all_lessons: all_lessons[current_key] = []
                all_lessons[current_key].append({
                    "id": f"{row_idx}", "question": question, "options": options, "correct": correct_options
                })
        return all_lessons
    except Exception as e:
        st.error(f"Lỗi: {e}")
        return None

# --- XỬ LÝ DỮ LIỆU ---
target_file = get_file_path()
data_by_lesson = load_all_data(target_file)
if not data_by_lesson: st.stop()

# --- SIDEBAR ---
st.sidebar.title("📂 Thư mục Bài học")
sorted_keys = sorted(data_by_lesson.keys(), key=natural_sort_key)
menu_display = []
for k in sorted_keys:
    score = st.session_state.high_scores.get(k, None)
    label = f"{k} (Điểm: {score})" if score is not None else k
    menu_display.append(label)

selected_label = st.sidebar.radio("Chọn bài học:", menu_display)
selected_lesson = selected_label.split(" (Điểm:")[0]

# --- NỘI DUNG ---
st.title(f"📖 {selected_lesson}")
is_submitted = selected_lesson in st.session_state.submitted_lessons
current_quiz = data_by_lesson[selected_lesson]
user_answers = {}

for item in current_quiz:
    st.subheader(f"Câu hỏi: {item['question']}")
    if len(item['correct']) > 1:
        st.caption("ℹ️ *Chọn nhiều đáp án*")
        selected = [opt for opt in item['options'] if st.checkbox(opt, key=f"c_{item['id']}_{opt}_{selected_lesson}", disabled=is_submitted)]
        user_answers[item['id']] = selected
    else:
        ans = st.radio("Chọn đáp án đúng:", item['options'], index=None, key=f"r_{item['id']}_{selected_lesson}", disabled=is_submitted)
        user_answers[item['id']] = ans

    if is_submitted:
        u_ans = user_answers.get(item['id'])
        is_correct = (set(u_ans) == set(item['correct']) and len(u_ans) == len(item['correct'])) if isinstance(u_ans, list) else (u_ans == item['correct'][0] if item['correct'] else False)
        if is_correct: st.success(f"🎯 Đúng! Đáp án: {', '.join(item['correct'])}")
        else: st.error(f"❌ Sai! Đáp án đúng: {', '.join(item['correct'])}")
    st.divider()

# --- CHẤM ĐIỂM VÀ HIỂN THỊ KẾT QUẢ ---
if not is_submitted:
    if st.button("Nộp bài", type="primary", use_container_width=True):
        score = 0
        for item in current_quiz:
            u_ans = user_answers.get(item['id'])
            c_ans = item['correct']
            
            # Logic chấm điểm chặt chẽ
            if isinstance(u_ans, list):
                if set(u_ans) == set(c_ans) and len(u_ans) == len(c_ans):
                    score += 1
            elif u_ans == (c_ans[0] if c_ans else None):
                score += 1
        
        # Cập nhật điểm cao nhất vào Sidebar
        if score > st.session_state.high_scores.get(selected_lesson, 0):
            st.session_state.high_scores[selected_lesson] = score
        
        st.session_state.submitted_lessons.add(selected_lesson)
        st.rerun()
else:
    # TÍNH LẠI ĐIỂM HIỆN TẠI ĐỂ HIỂN THỊ DÒNG THÔNG BÁO
    current_score = 0
    for item in current_quiz:
        u_ans = user_answers.get(item['id'])
        if isinstance(u_ans, list):
            if set(u_ans) == set(item['correct']) and len(u_ans) == len(item['correct']):
                current_score += 1
        elif u_ans == (item['correct'][0] if item['correct'] else None):
            current_score += 1

    # HIỂN THỊ DÒNG ĐIỂM ĐƠN GIẢN
    st.write("---")
    st.write(f"Kết quả: {current_score} / {len(current_quiz)} câu đúng.")
    
    if st.button("Làm lại bài tập"):
        st.session_state.submitted_lessons.remove(selected_lesson)
        st.rerun()