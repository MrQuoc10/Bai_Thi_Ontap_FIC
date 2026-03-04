import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os, re, random, time

st.set_page_config(page_title="Hệ Thống Ôn Tập FIC 2026", layout="wide")

# --- 1. CSS & UTILS ---
HIDE_SIDEBAR_CSS = "<style>[data-testid='stSidebar'] {display: none;} [data-testid='stSidebarNav'] {display: none;}</style>"

@st.cache_data
def load_all_data():
    files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('~$')]
    if not files: return None
    wb = load_workbook(files[0], data_only=True)
    ws = wb.active
    all_lessons = {}
    last_no, last_name = "", ""
    for r in range(1, ws.max_row + 1):
        val_a = ws.cell(row=r, column=1).value
        val_b = ws.cell(row=r, column=2).value
        if val_a and "Bài" in str(val_a):
            last_no = str(val_a).strip(); last_name = str(val_b).strip() if val_b else ""
        if not last_no: continue
        key = f"{last_no}: {last_name}"
        ques = ws.cell(row=r, column=4).value
        if not ques or str(ques).strip() in ["Câu hỏi", "STT"]: continue
        opts, corrects = [], []
        for c in range(6, 12):
            cell = ws.cell(row=r, column=c)
            if cell.value:
                val = str(cell.value).strip(); opts.append(val)
                # Nhận diện đáp án đúng qua màu nền
                if cell.fill and cell.fill.start_color and str(cell.fill.start_color.index) not in ("00000000", "FFFFFFFF", "0", "None"):
                    corrects.append(val)
        if opts:
            if key not in all_lessons: all_lessons[key] = []
            all_lessons[key].append({"id": f"{r}", "question": ques, "options": opts, "correct": corrects})
    return all_lessons

data = load_all_data()

# --- 2. STATE MANAGEMENT ---
if "page" not in st.session_state: st.session_state.page = "Luyện tập"
if "start_time" not in st.session_state: st.session_state.start_time = None
if "last_lesson" not in st.session_state and data: 
    st.session_state.last_lesson = sorted(data.keys())[0]

def reset_to_practice():
    st.session_state.page = "Luyện tập"
    st.session_state.start_time = None
    if "exam_list" in st.session_state: del st.session_state.exam_list
    st.rerun()

# --- 3. SIDEBAR ---
if st.session_state.page != "Thi tổng hợp":
    with st.sidebar:
        st.title("⭐ FIC 2026")
        if st.button("🔥 THI TỔNG HỢP (60 CÂU)", use_container_width=True, type="primary"):
            all_q = []
            for k in data: all_q.extend(data[k])
            st.session_state.exam_list = random.sample(all_q, min(60, len(all_q)))
            st.session_state.page = "Thi tổng hợp"
            st.session_state.start_time = time.time()
            st.rerun()
        st.divider()
        lesson_keys = sorted(data.keys(), key=lambda x: [int(t) if t.isdigit() else t.lower() for t in re.split('([0-9]+)', x)])
        selected_lesson = st.radio("Chọn bài:", lesson_keys, index=lesson_keys.index(st.session_state.last_lesson) if st.session_state.last_lesson in lesson_keys else 0)
        if st.session_state.last_lesson != selected_lesson:
            st.session_state.last_lesson = selected_lesson
            st.session_state.page = "Luyện tập"; st.rerun()

# --- 4. GIAO DIỆN CÂU HỎI ---
def render_question(item, mode="practice"):
    is_multiselect = len(item['correct']) > 1
    st.write(f"**{item['question']}**" + (" *(Chọn nhiều đáp án)*" if is_multiselect else ""))
    
    if mode == "practice":
        if is_multiselect:
            user_ans = []
            for opt in item['options']:
                if st.checkbox(opt, key=f"check_{item['id']}_{opt}"):
                    user_ans.append(opt)
            
            if st.button("Kiểm tra", key=f"btn_{item['id']}"):
                if set(user_ans) == set(item['correct']):
                    st.success("✅ Chính xác!")
                else:
                    st.error(f"❌ Sai. Đáp án đúng: {', '.join(item['correct'])}")
        else:
            ans = st.radio("Trả lời:", item['options'], index=None, key=f"prac_{item['id']}")
            if ans:
                if ans in item['correct']: st.success("✅ Chính xác!")
                else: st.error(f"❌ Sai. Đáp án đúng: {', '.join(item['correct'])}")
    else: # Mode Thi thử / Thi tổng hợp
        if is_multiselect:
            user_ans = []
            for opt in item['options']:
                if st.checkbox(opt, key=f"exam_{item['id']}_{opt}"):
                    user_ans.append(opt)
            return user_ans
        else:
            return st.radio("Chọn:", item['options'], index=None, key=f"exam_{item['id']}", label_visibility="collapsed")

# --- 5. LOGIC TRANG ---
if not data: st.error("Không tìm thấy dữ liệu!"); st.stop()

if st.session_state.page == "Thi tổng hợp":
    st.markdown(HIDE_SIDEBAR_CSS, unsafe_allow_html=True)
    st.title("🏆 Bài Thi Tổng Hợp (60 Câu)")
    elapsed = time.time() - st.session_state.start_time
    remaining = max(0, 3600 - int(elapsed))
    st.metric("Thời gian còn lại", f"{remaining//60:02d}:{remaining%60:02d}")

    ans_dict = {}
    for i, item in enumerate(st.session_state.exam_list):
        st.write(f"--- Câu {i+1} ---")
        ans_dict[item['id']] = render_question(item, mode="exam")
    
    st.divider()
    c1, c2 = st.columns([1, 5])
    if c1.button("🚪 Thoát bài", use_container_width=True): reset_to_practice()
    if c2.button("📤 Nộp bài thi tổng", type="primary", use_container_width=True) or remaining <= 0:
        correct = 0
        for item in st.session_state.exam_list:
            u_ans = ans_dict.get(item['id'])
            if isinstance(u_ans, list):
                if set(u_ans) == set(item['correct']): correct += 1
            elif u_ans == (item['correct'][0] if item['correct'] else None):
                correct += 1
        st.session_state.last_score = round((correct / len(st.session_state.exam_list)) * 100, 2)
        st.session_state.page = "Kết quả thi"; st.rerun()
    time.sleep(1); st.rerun(); st.stop()

elif st.session_state.page == "Thi thử bài":
    st.title(f"⏱️ Thi thử: {st.session_state.last_lesson}")
    ans_dict = {}
    current_q = data[st.session_state.last_lesson]
    for i, item in enumerate(current_q):
        st.write(f"--- Câu {i+1} ---")
        ans_dict[item['id']] = render_question(item, mode="exam")
    
    st.divider()
    c1, c2 = st.columns([1, 5])
    if c1.button("🚪 Hủy thi"): reset_to_practice()
    if c2.button("📤 Nộp bài thi thử", type="primary", use_container_width=True):
        correct = 0
        for item in current_q:
            u_ans = ans_dict.get(item['id'])
            if isinstance(u_ans, list):
                if set(u_ans) == set(item['correct']): correct += 1
            elif u_ans == (item['correct'][0] if item['correct'] else None):
                correct += 1
        st.session_state.last_score = round((correct / len(current_q)) * 100, 2)
        st.session_state.page = "Kết quả thi"; st.rerun()
    st.stop()

elif st.session_state.page == "Kết quả thi":
    st.title("🏁 Kết quả bài thi")
    st.header(f"Điểm của bạn: {st.session_state.last_score} / 100")
    if st.button("Quay lại Luyện tập"): reset_to_practice()
    st.stop()

else:
    st.title(f"📖 Luyện tập: {st.session_state.last_lesson}")
    if st.button(f"🚀 BẮT ĐẦU THI THỬ {st.session_state.last_lesson.split(':')[0]}", type="secondary"):
        st.session_state.page = "Thi thử bài"; st.rerun()
    st.divider()
    for item in data[st.session_state.last_lesson]:
        render_question(item, mode="practice")
        st.write("")
